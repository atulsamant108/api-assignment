
from collections import Counter

l1 = [11, 22, 48, 11, 98, 90, 22, 98, 11, 22, 22, 11]
item_count = Counter(l1)
print(dict(item_count))
#=========================================================

add = lambda x, y: x + y
print(add(3,7))

#===================================

'''
Instance method---> works on object data and is used for Selenium actions like click, type, get text.
Class method---> works on class data and is used for browser setup, config loading, and connections.
Static method----> is used for utilities like screenshots, waits, Excel reading, and data generation.
'''
class TestLogin:

    @classmethod
    def setup_class(cls):
        cls.driver = "Chrome"   # open browser
        print("Browser started")

    def test_login(self):
        print("Running test")

    @staticmethod
    def get_test_data():
        return ("admin", "1234")

#===================================
'''
A decorator in Python is a function that takes another function as input and adds extra functionality to it, 
without changing its code. Decorators are often used for logging, access control, timing, etc.'''

def my_decorator(func):
    def wrapper():
        print("Something before the function runs.")
        func()
        print("Something after the function runs.")
    return wrapper

@my_decorator
def say_hello():
    print("Hello!")

say_hello()
#===================================
'''
filter, map, and reduce are functional programming tools in Python:
filter---> Selects items from a sequence based on a condition.
map----> Applies a function to every item in a sequence.
reduce---> Applies a function cumulatively to items in a sequence, reducing it to a single value.
'''
from functools import reduce

numbers = [1, 2, 3, 4, 5]

# filter: keep even numbers
evens = list(filter(lambda x: x % 2 == 0, numbers))
print(evens)  # Output: [2, 4]

# map: square each number
squares = list(map(lambda x: x ** 2, numbers))
print(squares)  # Output: [1, 4, 9, 16, 25]

# reduce: sum all numbers
total = reduce(lambda x, y: x + y, numbers)
print(total)  # Output: 15
#===================================
'''
You expect specific exception during automation. How will you handle this scenario in pytest Python automation?
Ans--->When you expect an exception, Pytest provides a built-in context manager. pytest.raises()
Test passes if exception occurs
❌ Test fails if exception does NOT occur
'''
import pytest
from selenium.common.exceptions import NoSuchElementException

def test_missing_element(driver):
    with pytest.raises(NoSuchElementException):
        driver.find_element(By.ID, "non_existing_element")
#===================================
'''
What is Parameterization in PyTest (with Selenium)?
Ans--->Parameterization means running the same test case multiple times with different test data.
Instead of writing 5 separate tests for 5 inputs, you write one test and PyTest runs it many times with different values.
'''
#Without Parameterization ❌ You would normally write:
def test_login_admin(driver):
    driver .find_element(By.ID, "username").send_keys("admin")
    driver.find_element(By.ID, "password").send_keys("admin123")
    driver.find_element(By.ID, "login").click()
    assert  "Dashboard" in driver.title

def test_login_user(driver):
    driver.find_element("id", "username").send_keys("user")
    driver.find_element("id", "password").send_keys("user123")
    driver.find_element(By.ID, "login").click()
    assert "Dashboard" in driver.title

#With Parameterization ✅ You can write:
import pytest

@pytest.mark.parametrize("username, password", [ ("admin", "admin123"), ("user", "user123"), ("guest", "guest123") ] )
def test_login(driver, username, password):
    driver.find_element("id", "username").send_keys(username)
    driver.find_element("id", "password").send_keys(password)
    driver.find_element("id", "login").click()

    assert "Dashboard" in driver.title
 #If 1,000 test data rows exist → PyTest runs 1 test 1,000 times.hence its called 'Data-Driven Testing'
#==================================================================================================
'''
In the nexus to above question, how to read test data from Excel / CSV?
'''
# 1. Read data from CSV file, Example login_data.csv
'''username,password,expected
admin,admin123,success
user,user123,success
wrong,wrong123,fail'''

#below code will be in conftest.py or test file
import csv

def get_csv_data(file_path):
    data = []
    with open(file_path, newline='') as file:
        reader = csv.DictReader(file)
        for row in reader:
            data.append((row["username"], row["password"], row["expected"]))
    return data

#Use in PyTest
import pytest

@pytest.mark.parametrize("username,password,expected", get_csv_data("login_data.csv"))
def test_login(driver, username, password, expected):
    driver.find_element("id","username").send_keys(username)
    driver.find_element("id","password").send_keys(password)
    driver.find_element("id","login").click()

    if expected == "success":
        assert "Dashboard" in driver.title
    else:
        assert "Invalid" in driver.page_source

# 2. Read data from Excel (.xlsx)
#  run this command first --->pip install openpyxl
# Read Excel data
from openpyxl import load_workbook

def get_excel_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

#📌 Use in PyTest
import pytest

@pytest.mark.parametrize("username,password,expected", get_excel_data("login.xlsx"))
def test_login(driver, username, password, expected):
    driver.find_element("id","username").send_keys(username)
    driver.find_element("id","password").send_keys(password)
    driver.find_element("id","login").click()

    if expected == "success":
        assert "Dashboard" in driver.title
    else:
        assert "Invalid" in driver.page_source
#==================================================================================================
# Program to find duplicate elements in a list
numbers = [1, 2, 3, 4, 2, 5, 6, 3, 7, 8, 1]

duplicates = []
seen = set() #u can use list also but set is faster

for num in numbers:
    if num in seen:
        duplicates.append(num)
    else:
        seen.add(num)

print("Duplicate elements:", duplicates)
#==================================
# Find the closest pair of numbers in the list
l1 = [10, 44, 87, 23, 67, 90]

# Sort the list to make comparison easier
l1_sorted = sorted(l1)
min_diff = float('inf')
closest_pair = ()

for i in range(len(l1_sorted) - 1):
    diff = abs(l1_sorted[i+1] - l1_sorted[i])
    print(diff)
    if diff < min_diff:
        min_diff = diff
        closest_pair = (l1_sorted[i], l1_sorted[i+1])

print(f"Closest pair: {closest_pair} with difference {min_diff}")
#==================================
#print the first non-repeating character in a string
from collections import Counter

s = "aabbcde"
freq = Counter(s)

for ch in s:
    if freq[ch] == 1:
        print(ch)
        break

#==================================
#difference beteen while, continue and break
# Python
i = 0
while i < 5:         # while loop
    if i == 2:
        i += 1
        continue     # skip printing 2
    if i == 4:
        break        # exit loop when i is 4
    print(i)
    i += 1
# =============================================
#Q--->a =“atula” , separate vowels  {“a”:2, “u”:1}, write a python program for it
a = 'atula'
vowels = 'aeiou'
result = {}
for ch in a:
    if ch in vowels:
        result[ch] = result.get(ch,0) + 1
print(result)
#=============================================
"""
Q.1) There is a link on a page named Atul and contains text Atul, writer, xpath for it.
Ans---> //a[contains(text(), 'Atul')]
"""
# ==========================================================================================
# Q.2) There is a table on the UI having rows and columns. If you want to iterate through rows and columns, how will you do it

table = driver.find_element("xpath", "//table")

rows = table.find_elements("xpath", ".//tr")

for row in rows:
    columns = row.find_elements("xpath", ".//td")

    for col in columns:
        print(col.text)
# ==========================================================================================
#find second largest number in the list
numbers = [10, 5, 20, 8, 15]

largest = numbers[0]
second_largest = numbers[0]

for num in numbers:
    if num > largest:
        largest = num

for num in numbers:
    if second_largest < num < largest:
        second_largest = num

print(f"Largest number: {largest}")
print(f"Second largest number: {second_largest}")
#===================================
# Dictionary comprehension
squares = { x: x*x for x in range(5)} # dictionary comprehension, creates a dictionary where keys are numbers from 0 to 4 and values are their squares
print(squares)
# difference between get and [] in dictionary
print(squares[5]) # squares[5] # This will raise a KeyError because key 5 does not exist
print(squares.get(5)) # get method returns None if key is not found instead of raising an error
#===================================
"""
write a program to reverse a string without keeping spaces intact. For example, "Atul Samant" should become "tluA tnamaS".
This was asked in f2f interview in wipro
"""
name1 = "Atul Samant"
reversed_name = name1[::-1] # reverse the string using slicing
print(reversed_name) # Output: "tnamaS ltuA"
# If you want output as "lutA tnamaS" (reversed words but spaces in original position), you can do:
name2 = "Atul Samant"
reverse_name = " ".join(word[::-1] for word in name2.split(" "))
print(f"Original String: {name2}")
print(f"Reversed String: {reverse_name}")
#=============================================
#wite program
a = "Atul Samant"
output = "tnamaS ltuA"
#==============================
"""
Nums = [2433, 7567, 2458, 97760, 5654]

# Step 1: Sum of digits
digit_sums = list(map(lambda n: sum(map(int, str(n))), Nums))
print("Sum of digits:", digit_sums)

# Step 2: Divide each number by 2
quotients = list(map(lambda n: n/2, Nums))
print("Quotients:", quotients)

# Step 3: Numbers divisible by 3 (mapped to original list)
result = list(filter(lambda n: n % 3 == 0, Nums))

here in step2, I wanted the numbers of digit_sums which are divisible by 2. those numbers will be stored in quotients list then at last out of quotients list, whichever numbers are divisible by 3 for those numbers, the corresponding numbers from original list will be printed
print("Numbers divisible by 3:", result)
"""
Nums = [2433, 7567, 2458, 97760, 5654]

digit_sums = list(map(lambda n: sum(map(int, str(n))), Nums))

result = [Nums[i] for i in range(len(Nums))
          if digit_sums[i] % 2 == 0 and digit_sums[i] % 3 == 0]

print(result)

print("Corresponding numbers from original list:", result)
#=============================================
# program to find anagrams in a list of words
words = ["JAVA", "PQR", "LMN", "OPR", "VAJA", "RPQ"]

# Dictionary to store grouped anagrams
anagrams = {}

for word in words:
    # Sort the word to create a key
    key = ''.join(sorted(word))

    # Add word to corresponding key
    if key in anagrams:
        anagrams[key].append(word)
    else:
        anagrams[key] = [word]
"""
after  the output will be:
# {
    #     'AAJV': ['JAVA', 'VAJA'],
    #     'PQR': ['PQR', 'RPQ'],
    #     'LMN': ['LMN'],
    #     'OPR': ['OPR']
    # }
"""

# Print only groups that have more than one word (actual anagrams)
for group in anagrams.values():
    if len(group) > 1:
        print(group)

# The same above program can be written in a more concise way using defaultdict from collections module:
from collections import defaultdict
anagrams = defaultdict(list)

for word in words:
    key = ''.join(sorted(word))
    anagrams[key].append(word)

result = list(anagrams.values())

print(result)

#=======================================================
"""
1. Instance method uses self and accesses object variables
2. Class method uses cls and accesses class variables
3. Static method uses neither self nor cls and acts like a helper function inside the class. used for utility functions
that don't need access to instance or class data.

1.Instance Method → Student’s personal details (name, marks)
2. Class Method → School-level info (same for all students)
3. Static Method → General rule (like pass/fail logic) which doesn’t depend on any specific student or school.
"""

class Student:
    school_name = 'ABC School' # class variable

    def __init__(self, name, marks):
        self.name = name       # Instance variable
        self.marks = marks     # Instance variable

    def display_details(self):
        print("Name:", self.name)
        print("Marks:", self.marks)

    @classmethod
    def get_school_name(cls):
        print("School Name:", cls.school_name)

    @staticmethod
    def is_passed(marks):
        return marks >= 40

# creating object
student1 = Student("Atul", 85)

# Calling Instance Method
student1.display_details()

# Calling Class Method
Student.get_school_name()

# Calling Static Method
print("Pass Status:", Student.is_passed(75))
#===========================================================
employees = [
				{"name": "Alice", "salary": 88000},
				{"name": "Bob", "salary": 68000},
				{"name": "Charlie", "salary": 98000}
				]


max_salary_employee = max (employees, key=lambda x: x["salary"])
print (max_salary_employee["name"])
#============================================================
'''
Bharti Airtel interview questions--->
1. find a pair having the minimum time difference. ['00:00', '11:59', '12:00', '01:59'], explain the code properly
2. {
    "query": {
        "bool": {
            "must": [
                {
                    "match": {
                        "message": "73EA61FC92514D388280B885EAF8A7BD"
                    }
                }
            ],
            "filter": [
                {
                    "range": {
                        "@timestamp": {
                            "gte": "2025-01-02T12:32:45",
                            "lte": "2025-01-09T12:32:45"
                        }
                    }
                }
            ]
        }
    },
    "size": 100
}

a. write a code to print  "gte": "2025-01-02T12:32:45"
b. Find Count of keys whos value datatype is string for above json
3. in a flipcart website, if search something in searchbox, write a locator that will always select third option from searchbox
'''
#=====================Accionlabs - 11 May=================
'''
s1 = "ABCD", s2 = "CDAB" So here we need to check whether S2 is the rotation of S1 string, and if yes, then in which 
rotation of S1 we will get S2.
'''
s1 = "ABCD"
s2 = "CDAB"

found = False

for i in range(len(s1)):

    # Rotate string
    rotated = s1[i:] + s1[:i]

    # Check match
    if rotated == s2:
        print("S2 is rotation of S1")
        print("Rotation count:", i)
        found = True
        break

if not found:
    print("S2 is not rotation of S1")
#===============================================================================================